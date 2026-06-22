from dataclasses import fields, astuple

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from avito_parser.models import Item


def save_to_xlsx(items: list[Item], output_file: str) -> None:
    field_names = [f.name for f in fields(Item)]

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Avito"

    # Заголовки
    sheet.append(field_names)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Данные
    url_col_idx = field_names.index("url") + 1
    for item in items:
        sheet.append(astuple(item))

    # Форматирование ссылок
    for row in range(2, len(items) + 2):
        cell = sheet.cell(row=row, column=url_col_idx)
        cell.hyperlink = cell.value
        cell.font = Font(color="0563C1", underline="single")

    # Автоширина колонок
    for col_idx, name in enumerate(field_names, start=1):
        max_len = max([len(name)] + [len(str(getattr(item, name))) for item in items])
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Закрепление заголовков
    sheet.freeze_panes = "A2"
    wb.save(output_file)
